import openpyxl
import pandas
import subprocess
import platform
import os
import json
import shutil
import pandas as pd
from pandas import NaT
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

import time
import functools

from nadoo_telemarketing.CONSTANTS import KUNDENDATEN_ORDNER_APP, ARCHIV_ORDNER, BASE_DIR, UPDATE_ORDNER_NAME_USER, SETTINGS_ORDNER, UPDATE_ORDNER_NAME_APP

from nadoo_telemarketing.utils import (
    get_base_dir_path,
)
import json

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

def setup_folders():
    # Grund Verzeichnisstruktur anlegen
    ensure_base_folder_exits()
    ensure_folder_exists(UPDATE_ORDNER_NAME_USER)
    ensure_folder_exists(SETTINGS_ORDNER)
    

def get_settings_file_path():
    settings_folder = ensure_folder_exists(SETTINGS_ORDNER)
    settings_file = os.path.join(settings_folder, "settings.json")
    return settings_file

def get_settings():
    settings_file = get_settings_file_path()
    if os.path.exists(settings_file):
        with open(settings_file, "r") as file:
            return json.load(file)
    else:
        return {}

def set_settings(settings):
    settings_file = get_settings_file_path()
    with open(settings_file, "w") as file:
        json.dump(settings, file, indent=4)

def set_user_code(user_code):
    settings = get_settings()
    settings["user_code"] = user_code
    set_settings(settings)

def set_api_key(api_key):
    settings = get_settings()
    settings["api_key"] = api_key
    set_settings(settings)

def load_settings():
    settings_file = get_settings_file_path()
    if os.path.exists(settings_file):
        with open(settings_file, "r") as file:
            return json.load(file)
    else:
        return {}


def measure_execution_time(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        execution_time = end_time - start_time
        #print(f"Die Funktion '{func.__name__}' hat {execution_time:.2f} Sekunden gedauert.")
        return result
    return wrapper

""" 
Vorlage for JSON-Datei anbindung
def set_xyz_data(beweismittel_data):
    ensure_beweismittel_OWi_data_file_exists()
    file_path = get_beweismittel_OWi_data_file_path()
    with open(file_path, "w") as f:
        json.dump(beweismittel_data, f, indent=4)


def get_xyz_data():
    ensure_beweismittel_OWi_data_file_exists()
    file_path = get_beweismittel_OWi_data_file_path()
    with open(file_path, "r") as f:
        return json.load(f)
"""
def versionsordner_erstellen():
        folder_name = UPDATE_ORDNER_NAME_USER
        ensure_folder_exists(folder_name)
        
def ensure_folder_exists(folder_name):
    base_dir = get_base_dir_path()
    folder_path = os.path.join(base_dir, folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

def ensure_base_folder_exits():
    base_dir = os.path.join(os.path.expanduser("~"), BASE_DIR)
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)

def open_file(file_path):
    if file_path:
        try:
            if platform.system() == "Windows":
                subprocess.run(["explorer", file_path], check=True)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", file_path], check=True)
            else:  # Assuming Linux
                subprocess.run(["xdg-open", file_path], check=True)
        except subprocess.CalledProcessError as e:
            print(f"Error opening file: {e}")
    else:
        print("File path is not set. Unable to open file.")
        
def get_updates_datei_user():
    #ich möchte den dateipfad der update.json datei aus dem updates ordner haben, welcher im updates ordner im user verzeichnis liegt
    updates_dateipfad_user = os.path.join(os.path.expanduser("~"), get_base_dir(), UPDATE_ORDNER_NAME_USER, "update.json")
    return updates_dateipfad_user

def get_updates_pdf_path(app):
    updates_pdf_dateipfad = os.path.join(app.paths.app, "resources", "update.pdf")
    return updates_pdf_dateipfad

def update_daten_laden_user():
        file_path = get_updates_datei_user()
        with open(file_path, "r") as f:
            return json.load(f)
        
def update_daten_laden_app(app):
        file_path = get_updates_datei_app(app)
        with open(file_path, "r") as f:
            return json.load(f)
        
def get_updates_datei_app(app):
    updates_datei_app = os.path.join(app.paths.app, "resources", UPDATE_ORDNER_NAME_APP, "update.json")
    return updates_datei_app

def update_in_updates_ordner_uebertragen(app):
        updates_datei_app = os.path.join(app.paths.app, "resources", UPDATE_ORDNER_NAME_APP, "update.json")
        updates_datei_user = get_updates_datei_user()

        shutil.copy(updates_datei_app, updates_datei_user)
        
def get_help_file_path(app):
    return os.path.join(app.paths.app, "resources", "help.pdf")

def get_base_dir():
    return get_base_dir_path()

"""
#TODO #96 Veralgemeinern
def update_daten_in_basis_ordner_uebertragen(app):
    kundendaten_ordner_app = os.path.join(app.paths.app, "resources", KUNDENDATEN_ORDNER_APP)
    basis_ordner_user = get_base_dir()  # Angenommen, diese Funktion gibt den Basisordner des Benutzers zurück
    archiv_ordner = os.path.join(basis_ordner_user,ARCHIV_ORDNER, "archivierte_anwalt_daten")
    ziel_json_datei_name = "lawyer_details.json"
    update_datei_name = "update_lawyer_details.json"

    ziel_datei_pfad = os.path.join(basis_ordner_user, ziel_json_datei_name)
    update_datei_pfad = os.path.join(kundendaten_ordner_app, update_datei_name)

    # Stelle sicher, dass der Archivordner existiert
    if not os.path.exists(archiv_ordner):
        os.makedirs(archiv_ordner)

    # Überprüfe, ob die Datei lawyer_details.json bereits im Zielordner existiert
    if os.path.isfile(ziel_datei_pfad):
        # Wenn ja, und update_lawyer_details.json ist vorhanden, archiviere die alte Datei
        if os.path.isfile(update_datei_pfad):
            # Generiere einen eindeutigen Namen für die archivierte Datei
            datumsanhang = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            archivierte_datei_name = f"lawyer_details_{datumsanhang}.json"
            archivierte_datei_pfad = os.path.join(archiv_ordner, archivierte_datei_name)

            # Verschiebe die alte lawyer_details.json in den Archivordner
            shutil.move(ziel_datei_pfad, archivierte_datei_pfad)

            # Kopiere die update_lawyer_details.json in den Zielordner und benenne sie um
            shutil.copy(update_datei_pfad, ziel_datei_pfad)

    else:
        # Wenn keine lawyer_details.json im Zielordner, kopiere diese aus dem Kundendatenordner, falls vorhanden
        source_datei_pfad = os.path.join(kundendaten_ordner_app, ziel_json_datei_name)
        if os.path.isfile(source_datei_pfad):
            shutil.copy(source_datei_pfad, ziel_datei_pfad)
            
def vorlagen_in_vorlagen_ordner_uebertragen(app):
    vorlagen_ordner_app = os.path.join(app.paths.app, "resources", KUNDENDATEN_ORDNER_APP, VORLAGEN_ORDNER_APP)
    vorlagen_ordner_user = get_template_folder()  # Annahme, dass diese Funktion das Benutzerverzeichnis für Vorlagen zurückgibt
    archiv_ordner = os.path.join(vorlagen_ordner_user, ARCHIV_ORDNER)
    placeholder_file_name = "placeholder.txt"

    # Stelle sicher, dass der Archivordner existiert
    if not os.path.exists(archiv_ordner):
        os.makedirs(archiv_ordner)

    # Überprüfe, ob der Vorlagenordner existiert
    if not os.path.isdir(vorlagen_ordner_app):
        return  # Beende die Funktion, wenn der Ordner nicht existiert

    for file in os.listdir(vorlagen_ordner_app):
        if file == placeholder_file_name:
            continue  # Ignoriere die Platzhalterdatei

        vorlage_app_pfad = os.path.join(vorlagen_ordner_app, file)
        neuer_name_ohne_update = file.replace("update_", "")
        vorlage_user_pfad = os.path.join(vorlagen_ordner_user, neuer_name_ohne_update)

        # Wenn es sich um eine Update-Datei handelt ODER keine Datei im Benutzerverzeichnis existiert, führe den Kopiervorgang durch
        if file.startswith("update_") or not os.path.exists(vorlage_user_pfad):
            if os.path.exists(vorlage_user_pfad):
                # Archiviere die existierende Datei, bevor sie durch die Update-Datei ersetzt wird
                basisname, erweiterung = os.path.splitext(neuer_name_ohne_update)
                datumsanhang = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                archivierte_datei_name = f"{basisname}_{datumsanhang}{erweiterung}"
                shutil.move(vorlage_user_pfad, os.path.join(archiv_ordner, archivierte_datei_name))
            
            # Kopiere die Vorlage oder aktualisierte Vorlage in das Benutzerverzeichnis
            shutil.copy(vorlage_app_pfad, vorlage_user_pfad)

            if file.startswith("update_"):
                # Umbenennen der "update_"-Datei im Anwendungsordner, indem das "update_" Präfix entfernt wird
                neuer_app_pfad_ohne_update = os.path.join(vorlagen_ordner_app, neuer_name_ohne_update)
                os.rename(vorlage_app_pfad, neuer_app_pfad_ohne_update)
"""
   
def finde_letzte_zeile_in_excel_blatt(sheet:Worksheet):
                
    # finde die letzte Spalte mit Werten
    anzahl_der_spalten = 1
    
    while (
        sheet.cell(1, anzahl_der_spalten).value is not None
        and sheet.cell(1, anzahl_der_spalten).value != "Spalte1"
    ):
        anzahl_der_spalten += 1
                
    max_rows = 1
     
    vollstaendig_leere_zeile_gefunden = False
    leere_spalten_gefunden = 0
    spalte = 1
    while spalte <= anzahl_der_spalten and not vollstaendig_leere_zeile_gefunden:

        while sheet.cell(max_rows, spalte).value is not None:
            max_rows += 1

        if sheet.cell(max_rows, spalte).value is None:
            #print("Leere Spalte gefunden")
            leere_spalten_gefunden += 1
        else:
            leere_spalten_gefunden = 0

        spalte += 1
        
        if leere_spalten_gefunden == anzahl_der_spalten:
            vollstaendig_leere_zeile_gefunden = True
        elif spalte > anzahl_der_spalten:
            spalte = 1
            
    max_rows = max_rows - 1        
    return max_rows
    


def uebertrage_daten_in_excel_blatt(sheet: Worksheet, daten: pd.DataFrame):
    max_rows = finde_letzte_zeile_in_excel_blatt(sheet)
    """
    print(
        f"Die Tabelle in der Region von Zeile 1 bis {max_rows} und Spalte 1 bis {len(daten.columns)} hat Daten"
    )
    """

    # Übertrage die Daten aus dem DataFrame in das Excel-Blatt unterhalb der vorhandenen Daten
    for row_num, row_data in enumerate(daten.values, start=max_rows + 1):
        for col_num, cell_value in enumerate(row_data, start=1):
            if col_num == 1:
                # Index wird übersprungen
                continue
            col_letter = get_column_letter(col_num-1)
            
            # Überprüfe, ob der Zellwert ein Datum ist und formatiere es entsprechend
            if isinstance(cell_value, datetime):
                if pd.isna(cell_value):  # Überprüfung, ob der Wert NaT ist
                    cell_value = ''  # Oder ein anderer Platzhalterwert deiner Wahl
                else:
                    cell_value = cell_value.strftime("%d.%m.%Y")
            
            sheet[f"{col_letter}{row_num}"] = cell_value
def zeilentrennung_am_ende_des_excel_blatt_anfuegen(sheet: Worksheet):
    max_rows = finde_letzte_zeile_in_excel_blatt(sheet)
    if max_rows > 1:
        spaltenbezeichner_status_column = finde_spaltenbezeichner_zahlenwert(sheet, "Status")
        print(f"Spaltenbezeichner Status: {spaltenbezeichner_status_column}")
        # Füge eine Zeile mit - in alle Zellen der letzten Reihe hinzu und färbe den Hintergrund rot
        for col_num in range(1, spaltenbezeichner_status_column+1):
            print(f"Spalte {col_num}")
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{max_rows+1}"] = "-"
            sheet[f"{col_letter}{max_rows+1}"].fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )


def finde_spaltenbezeichner_zahlenwert(excel_blatt, spaltenbezeichner):
    """
    Findet den Zahlenwert für den gegebenen Spaltenbezeichner in einem Excel-Blatt.

    :param excel_blatt: Das Excel-Blatt, in dem nach dem Spaltenbezeichner gesucht werden soll.
    :param spaltenbezeichner: Der Spaltenbezeichner (z.B. "Status"), nach dem gesucht werden soll.
    :return: Der Zahlenwert des Spaltenbezeichners (z.B. 1 für Spalte A, 2 für Spalte B usw.) oder None, wenn der Spaltenbezeichner nicht gefunden wurde.
    """
    # Entferne Leerzeichen am Ende der Spaltenüberschriften
    headers = [cell.value.strip() if cell.value is not None else cell.value for cell in excel_blatt[1]]

    # Finde den Zahlenwert des Spaltenbezeichners basierend auf den bereinigten Spaltenüberschriften
    for i, header in enumerate(headers, start=1):
        if header == spaltenbezeichner:
            return i

    # Wenn der Spaltenbezeichner nicht gefunden wurde, gib None zurück
    return None
    
def lade_daten_aus_excel_blatt_als_DataFrame(excel_blatt: Worksheet, spaltenbezeichner_status: str = "Status") -> pd.DataFrame:
    """
    Lädt alle Daten aus einem Excel-Blatt in ein DataFrame.

    Args:
        excel_blatt (openpyxl.worksheet.worksheet.Worksheet): Das Excel-Blatt, aus dem die Daten geladen werden sollen.
        spaltenbezeichner_status (str, optional): Der Name des Spaltenbezeichners für den Status. Standard ist "Status".

    Returns:
        pd.DataFrame: Ein DataFrame mit allen Daten aus dem Excel-Blatt.
    """
    # Finde den Spaltenbezeichner (A1, B1, C1 ...) mit dem Namen Status
    spaltenbezeichner_status_column = None

    # Entferne Leerzeichen am Ende der Spaltenüberschriften
    headers = [cell.value.strip() if cell.value is not None else cell.value for cell in excel_blatt[1]]

    # Finde den Spaltenbezeichner basierend auf den bereinigten Spaltenüberschriften
    spaltenbezeichner_status_column = finde_spaltenbezeichner_zahlenwert(excel_blatt, spaltenbezeichner_status)

    # Finde die letzte Zeile im Excel-Blatt
    max_rows = finde_letzte_zeile_in_excel_blatt(excel_blatt)
    print(max_rows)

    # Lade die Daten aus dem Excel-Blatt in ein DataFrame
    data = list(excel_blatt.iter_rows(min_row=2, max_row=max_rows, max_col=spaltenbezeichner_status_column, values_only=True))
    dataframe_aller_daten = pd.DataFrame(data, columns=headers[:spaltenbezeichner_status_column])

    return dataframe_aller_daten


def loesche_alle_daten_von_blatt(excel_blatt: Worksheet):
    """
    Löscht alle Daten von einem Excel-Blatt.

    Args:
        excel_blatt: Das Excel-Blatt, aus dem die Daten gelöscht werden sollen.
        spaltenbezeichner_status (str, optional): Der Name des Spaltenbezeichners für den Status. Standard ist "Status".
    """

    # Finde die letzte Zeile im Excel-Blatt
    max_rows = finde_letzte_zeile_in_excel_blatt(excel_blatt)

    # Lösche alle Daten aus dem Excel-Blatt
    excel_blatt.delete_rows(2, max_rows)
    
def ermittel_den_aktuellen_monat_als_deutsches_wort(monat_offset=0):
    """
    Ermittelt den Monat relativ zum aktuellen Monat als Zahl und gibt den entsprechenden deutschen Monatsnamen zurück.

    Args:
        monat_offset (int, optional): Die Anzahl der Monate, die zum aktuellen Monat addiert werden sollen. Standard ist 0.

    Returns:
        str: Der ermittelte Monat auf Deutsch (z.B. "Januar", "Februar", ...).
    """
    # Erstelle ein Dictionary, das die Monatszahlen auf die deutschen Monatsnamen abbildet
    monatsnamen = {
        1: "Januar",
        2: "Februar",
        3: "März",
        4: "April",
        5: "Mai",
        6: "Juni",
        7: "Juli",
        8: "August",
        9: "September",
        10: "Oktober",
        11: "November",
        12: "Dezember"
    }

    # Ermittle den aktuellen Monat als Zahl
    aktueller_monat_als_zahl = datetime.now().month

    # Addiere den Offset zum aktuellen Monat
    ziel_monat = (aktueller_monat_als_zahl + monat_offset) % 12
    if ziel_monat == 0:
        ziel_monat = 12

    # Gib den entsprechenden deutschen Monatsnamen zurück
    return monatsnamen[ziel_monat]